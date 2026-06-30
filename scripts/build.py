"""
报销 Excel 生成工具
用法: python build.py <工作目录>

职责：读数据 → 配 PDF → 嵌图片 → 检查 → 输出 Excel
Claude 只需准备 数据.csv、payment_mapping.json、scan_mapping.json，
不需改动本脚本。
"""
import csv, io, json, os, re, sys
from copy import copy
from pathlib import Path

import pypdfium2 as pdfium
from PIL import Image
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU

# ═══════════════════════════════════════════════════════════════════════
# 配置
# ═══════════════════════════════════════════════════════════════════════

WORK = Path(sys.argv[1]).resolve()
SKILL_DIR = Path(__file__).resolve().parent
TEMPLATE = SKILL_DIR / '模板.xlsx'
DPI = 600
DISP_SINGLE = 200          # 单 PDF 显示宽度
DISP_PAIR   = 155          # 双 PDF 显示宽度
DISP_EXTRA  = 135          # 付款截图 / 扫描件显示宽度
GAP = 10

# 文件材料要求以 references/document-checklist.md 为单一事实来源。
# 此字典仅用于运行时自检，修改材料要求时请同步更新 document-checklist.md。
EXPECTED = {
    '高铁': {'发票': 1, '扫描件': 1, '付款截图': 1},
    '住宿': {'发票': 1, '扫描件': 1, '付款截图': 1},
    '打车': {'发票': 1, '行程单': 1, '扫描件': 1, '付款截图': 1},
    '其他': {'发票': 1, '扫描件': 1, '付款截图': 1},
}

# ═══════════════════════════════════════════════════════════════════════
# 1. 加载数据
# ═══════════════════════════════════════════════════════════════════════

def load_csv():
    p = WORK / '数据.csv'
    if not p.exists():
        die(f'{p} 不存在')
    rows = []
    with open(p, encoding='utf-8-sig') as f:
        rdr = csv.reader(f)
        for linenum, r in enumerate(rdr, start=1):
            if not r or not any(r):
                continue
            if len(r) != 13:
                die(f'数据.csv 第 {linenum} 行有 {len(r)} 列（应为 13 列），请检查 CSV 文件')
            try:
                r[5] = float(r[5])
            except (ValueError, IndexError):
                die(f'数据.csv 第 {linenum} 行金额列（第6列）无法转为数字: "{r[5]}"')
            if not re.match(r'^\d{4}-\d{2}$', str(r[6])):
                die(f'数据.csv 第 {linenum} 行月份格式错误 "{r[6]}"（应为 YYYY-MM）')
            if not str(r[12]).strip():
                die(f'数据.csv 第 {linenum} 行排序键（第13列）为空，每行必须填写排序键')
            if r[9] == '':
                r[9] = None
            rows.append(r)
    if not rows:
        die('数据.csv 没有有效数据行')
    rows.sort(key=lambda r: r[12])
    return rows

# ═══════════════════════════════════════════════════════════════════════
# 2. PDF 自动匹配
# ═══════════════════════════════════════════════════════════════════════

def match_pdfs(rows):
    # TODO: 后续应将 PDF 标识从文件名升级为相对路径，避免子目录同名文件碰撞。
    global pdf_paths
    pdfs = sorted(Path(WORK).rglob('*.pdf'))
    pdf_paths = {p.name: p for p in pdfs}
    mapping = {}
    assigned = set()

    # ── 高铁票：纯数字文件名 → K 列票号 ──
    for pdf in pdfs:
        m = re.match(r'^(\d+)\.pdf$', pdf.name)
        if m:
            ticket = m.group(1)
            for i, rd in enumerate(rows):
                if rd[10] == ticket:
                    mapping[3 + i] = [pdf.name]
                    assigned.add(pdf.name)
                    break

    # ── 住宿 dzfp_ → K 列发票号 ──
    for pdf in pdfs:
        if pdf.name.startswith('dzfp_') and pdf.name not in assigned:
            m = re.match(r'dzfp_(\d+)', pdf.name)
            if m:
                invoice = m.group(1)
                for i, rd in enumerate(rows):
                    if rd[10] == invoice:
                        mapping[3 + i] = [pdf.name]
                        assigned.add(pdf.name)
                        break

    # ── 高德聚合（先于滴滴，因为文件名含精确平台+金额）──
    gaode_inv  = [p for p in pdfs if '高德打车电子发票' in p.name and p.name not in assigned]
    gaode_trv  = [p for p in pdfs if '高德打车电子行程单' in p.name and p.name not in assigned]
    PLATFORMS = ['T3', '如祺', '风韵', '曹操', '享道']

    for inv in gaode_inv:
        plat = next((k for k in PLATFORMS if k in inv.name), None)
        trv = next((t for t in gaode_trv if plat and plat in t.name and t.name not in assigned), None)
        m_amt = re.search(r'(\d+\.\d{2})', inv.name)
        if m_amt and plat:
            amt = float(m_amt.group(1))
            for i, rd in enumerate(rows):
                r = 3 + i
                if r in mapping:
                    continue
                if abs(float(rd[5]) - amt) < 2 and plat in rd[11]:
                    mapping[r] = [inv.name] + ([trv.name] if trv else [])
                    assigned.update([inv.name] + ([trv.name] if trv else []))
                    break

    # ── 滴滴：无法通过文件名确定归属，留待 write_pdf_review_if_needed 处理 ──
    unmatched = [p.name for p in pdfs if p.name not in assigned]
    unmapped  = [3 + i for i in range(len(rows)) if (3 + i) not in mapping]
    return mapping, unmatched, unmapped

# ═══════════════════════════════════════════════════════════════════════
# 3. PDF 状态计算与人工映射校验
# ═══════════════════════════════════════════════════════════════════════

def valid_data_rows(rows):
    """返回有效 Excel 数据行号集合，例如 {3, 4, 5, ...}"""
    return {3 + i for i in range(len(rows))}


def flatten_pdf_map(pdf_map):
    """返回 pdf_map 中已经使用的 PDF 文件名集合"""
    return {pf for v in pdf_map.values() if v for pf in v}


def recompute_pdf_status(rows, pdf_map):
    """基于合并后的 pdf_map 重新计算 unmatched_pdfs 和 unmapped_rows"""
    all_pdfs = {p.name for p in Path(WORK).rglob('*.pdf')}
    in_use = flatten_pdf_map(pdf_map)
    unmatched = sorted(all_pdfs - in_use)
    unmapped = sorted(r for r in valid_data_rows(rows) if not pdf_map.get(r))
    return unmatched, unmapped


def load_pdf_mapping(name, valid_rows, all_pdf_paths, auto_pdf_map=None):
    """读取并校验 pdf_mapping.json。返回 (manual_map, errors, warnings)"""
    p = WORK / name
    if not p.exists():
        return {}, [], []

    errors = []
    warnings = []

    try:
        with open(p, encoding='utf-8') as f:
            raw = json.load(f)
    except json.JSONDecodeError as e:
        errors.append(f'{name} 格式错误: {e}')
        return {}, errors, warnings

    if not isinstance(raw, dict):
        errors.append(f'{name} 必须是 JSON 对象')
        return {}, errors, warnings

    # 收集所有文件中引用的 PDF 用于重复检测
    seen_pdfs = {}  # filename -> row

    manual_map = {}
    for key, val in raw.items():
        # 1. row key 校验
        try:
            row = int(key)
        except (ValueError, TypeError):
            errors.append(f'{name} 行号 "{key}" 不是有效数字')
            continue
        if row not in valid_rows:
            errors.append(f'{name} 行号 R{row} 不在有效数据行范围内（有效范围: R{min(valid_rows)}-R{max(valid_rows)}）')
            continue

        # 2. value 类型校验
        if not isinstance(val, list):
            errors.append(f'{name} R{row} 的值必须是数组，当前类型: {type(val).__name__}')
            continue

        # 3. 文件存在校验
        for fn in val:
            if fn not in all_pdf_paths:
                errors.append(f'{name} R{row} 引用了不存在的 PDF: {fn}')

        # 4. 重复引用校验
        for fn in val:
            if fn in seen_pdfs and seen_pdfs[fn] != row:
                errors.append(f'PDF 被重复映射到多个行: {fn} → R{seen_pdfs[fn]}, R{row}')
            else:
                seen_pdfs[fn] = row

        # 5. 空数组处理
        if not val:
            warnings.append(f'{name} R{row} 是空数组，不计入 PDF 映射')

        manual_map[row] = list(val)

    if errors:
        return manual_map, errors, warnings

    # 6. 人工映射 vs 自动映射冲突检测（同一 PDF 被分配到不同行）
    if auto_pdf_map:
        auto_pdf_owner = {}
        for auto_row, files in (auto_pdf_map or {}).items():
            for fn in files:
                auto_pdf_owner[fn] = auto_row
        for manual_row, files in manual_map.items():
            for fn in files:
                if fn in auto_pdf_owner and auto_pdf_owner[fn] != manual_row:
                    errors.append(
                        f'{name} R{manual_row} 引用了已自动映射到 R{auto_pdf_owner[fn]} 的 PDF: {fn}')

    # 7. 人工覆盖自动映射时提示（同行的不同值才报 warning）
    if errors:
        return manual_map, errors, warnings

    if auto_pdf_map:
        for row in manual_map:
            if manual_map[row] and row in auto_pdf_map and auto_pdf_map[row]:
                if manual_map[row] != auto_pdf_map[row]:
                    warnings.append(
                        f'{name} 覆盖自动 PDF 映射 R{row}'
                        f'  自动: {auto_pdf_map[row]}'
                        f'  人工: {manual_map[row]}')

    return manual_map, errors, warnings


def write_pdf_review_if_needed(rows, pdf_map):
    """根据当前 PDF 状态刷新 pdf_mapping_review.json。
    只包含当前仍未自动匹配、且未被人工映射解决的待确认 PDF。"""
    all_pdfs = {p.name for p in Path(WORK).rglob('*.pdf')}
    in_use = flatten_pdf_map(pdf_map)
    still_unmatched = sorted(all_pdfs - in_use)

    # 滴滴相关 PDF
    didi_inv  = sorted([fn for fn in still_unmatched if '滴滴电子发票' in fn])
    didi_trv  = sorted([fn for fn in still_unmatched if '滴滴出行行程报销单' in fn])
    # 高德相关（也属于无法自动匹配的聚合打车类）
    gaode_pdfs = sorted([fn for fn in still_unmatched if '高德' in fn])
    other_unmatched = sorted([fn for fn in still_unmatched
                             if '滴滴' not in fn and '高德' not in fn])

    # 仍然没有 PDF 的行中，打车相关的行
    unmapped_taxi_rows = []
    for i in range(len(rows)):
        r = 3 + i
        if r in pdf_map and pdf_map[r]:
            continue
        desc = rows[i][11]
        if any(k in desc for k in ['滴滴', '打车', '出行', 'T3', '如祺', '风韵', '曹操', '享道', '高德']):
            unmapped_taxi_rows.append({"row": r, "amount": rows[i][5], "desc": desc})

    review = {
        "_note": "当前仍有以下 PDF 无法自动确定归属行，请人工确认后写入 pdf_mapping.json。已写入 pdf_mapping.json 的 PDF 不会出现在此清单中。",
        "pending_didi_invoices": didi_inv,
        "pending_didi_travels": didi_trv,
        "pending_gaode_pdfs": gaode_pdfs,
        "other_unmatched_pdfs": other_unmatched,
        "unmapped_taxi_rows": unmapped_taxi_rows,
    }

    review_path = WORK / 'pdf_mapping_review.json'
    if didi_inv or didi_trv or gaode_pdfs:
        with open(review_path, 'w', encoding='utf-8') as f:
            json.dump(review, f, ensure_ascii=False, indent=2)
        print(f'  → 已刷新 {review_path}')
    elif review_path.exists():
        # 没有待确认项时写入空状态
        review["_note"] = "当前没有待人工确认的 PDF"
        with open(review_path, 'w', encoding='utf-8') as f:
            json.dump(review, f, ensure_ascii=False, indent=2)
        print(f'  ✓ 没有待确认 PDF，{review_path} 已清空')
    return review

# ═══════════════════════════════════════════════════════════════════════
# 4. 填充数据
# ═══════════════════════════════════════════════════════════════════════

def fill_sheet(ws, rows):
    ref_row = 7
    for i, rd in enumerate(rows):
        r = 3 + i
        for j, val in enumerate(rd[:12]):
            src = ws.cell(row=ref_row, column=j + 1)
            dst = ws.cell(row=r, column=j + 1)
            dst.value = val
            if src.has_style:
                dst.font = copy(src.font)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)
                dst.number_format = copy(src.number_format)
                dst.alignment = copy(src.alignment)

    n = len(rows)
    s, b, p = 3 + n, 4 + n, 5 + n  # sum, backup, pay rows

    ws.cell(row=s, column=1).value = '合计'
    ws.cell(row=s, column=6).value = f'=SUM(F3:F{2 + n})'
    ws.cell(row=s, column=7).value = '/'
    ws.cell(row=s, column=12).value = '/'
    ws.cell(row=b, column=1).value = '本次核销备用金金额'
    ws.cell(row=b, column=7).value = '之前领用过备用金的，本次发生的费用冲抵备用金的金额'
    ws.cell(row=p, column=1).value = '核销完备用金后需支付金额'
    ws.cell(row=p, column=6).value = f'=F{s}-F{b}'
    ws.cell(row=p, column=7).value = '/'

    for a, b_col, c, d_col in [(s, 'E', s, 'K'), (b, 'E', b, 'K'), (p, 'E', p, 'K')]:
        ws.merge_cells(f'A{a}:{b_col}{a}')
        ws.merge_cells(f'G{a}:{d_col}{a}')
    ws.merge_cells(f'L{s}:L{p}')

    for r in range(3, 3 + n):
        c = ws.cell(row=r, column=6)
        if c.value is not None and not str(c.value).startswith('='):
            c.number_format = '0.00'
    for r in range(p + 1, 50):
        for c in range(1, 13):
            ws.cell(row=r, column=c).value = None

    return s, b, p

# ═══════════════════════════════════════════════════════════════════════
# 5. 嵌入图片
# ═══════════════════════════════════════════════════════════════════════

# ── PDF lookup (populated in main) ──
pdf_paths = {}

def render_pdf(name):
    path = pdf_paths.get(name, WORK / name)
    doc = pdfium.PdfDocument(str(path))
    bmp = doc[0].render(scale=DPI / 72)
    img = bmp.to_pil()
    doc.close()
    return img

def embed_images(ws, pdf_map, pay_map, scan_map):
    pay_dir = WORK / '付款截图'
    scan_dir = WORK / '扫描件'

    for row_num in sorted(set(pdf_map) | set(pay_map) | set(scan_map)):
        pdfs  = pdf_map.get(row_num, [])
        pays  = pay_map.get(row_num, [])
        scans = scan_map.get(row_num, [])

        items = []
        if len(pdfs) == 1:
            items.append((render_pdf(pdfs[0]), DISP_SINGLE))
        elif len(pdfs) >= 2:
            for pf in pdfs:
                items.append((render_pdf(pf), DISP_PAIR))
        for pf in pays:
            items.append((Image.open(str(pay_dir / pf)), DISP_EXTRA))
        for sf in scans:
            items.append((Image.open(str(scan_dir / sf)), DISP_EXTRA))

        if not items:
            continue

        x_off, max_dh = 0, 0
        for img, dw in items:
            dh = int(img.height * dw / img.width)
            max_dh = max(max_dh, dh)
            buf = io.BytesIO()
            if img.mode == 'RGBA':
                img = img.convert('RGB')
            img.save(buf, format='JPEG', quality=92)
            buf.seek(0)
            xl = XLImage(buf)
            xl.width = dw
            xl.height = dh
            if x_off == 0:
                xl.anchor = f'J{row_num}'
            else:
                anchor = OneCellAnchor(
                    _from=AnchorMarker(col=9, colOff=pixels_to_EMU(x_off), row=row_num - 1, rowOff=0),
                    ext=XDRPositiveSize2D(pixels_to_EMU(dw), pixels_to_EMU(dh)))
                xl.anchor = anchor
            ws.add_image(xl)
            x_off += dw + GAP

        ws.row_dimensions[row_num].height = max_dh * 0.75 + 25
        print(f'  R{row_num}: {"+".join(str(dw) for _,dw in items)} h={max_dh}')

# ═══════════════════════════════════════════════════════════════════════
# 6. 文档完整性检查
# ═══════════════════════════════════════════════════════════════════════

def classify(rd):
    cat, desc = rd[8], rd[11]
    if cat == '住宿费':
        return '住宿'
    if any(k in desc for k in ['高铁', '二等座']):
        return '高铁'
    if any(k in desc for k in ['打车', '出行', '滴滴', 'T3', '如祺', '风韵', '曹操', '享道', '高德']):
        return '打车'
    return '其他'

def pdf_kind(fname):
    if '电子发票' in fname or fname.startswith('dzfp_'):
        return '发票'
    if '行程单' in fname or '行程报销单' in fname:
        return '行程单'
    return '发票'

def check_completeness(rows, pdf_map, pay_map, scan_map):
    print('\n── 文档完整性检查 ──')
    header = f'  {"行":4s} {"金额":>7s} {"项目":24s} {"类型":4s} {"发票":5s} {"行程单":6s} {"付款截图":7s} {"扫描件":6s}'
    print(header)
    print('  ' + '-' * (len(header) - 2))
    for i, rd in enumerate(rows):
        r = 3 + i
        ctype = classify(rd)
        exp = EXPECTED[ctype]

        present = {}
        for pf in pdf_map.get(r, []):
            k = pdf_kind(pf)
            present[k] = present.get(k, 0) + 1
        present['付款截图'] = present.get('付款截图', 0) + len(pay_map.get(r, []))
        present['扫描件'] = present.get('扫描件', 0) + len(scan_map.get(r, []))

        def cell(k):
            want = exp.get(k, 0)
            if want == 0:
                return '  —  '
            got = present.get(k, 0)
            return '  ✓  ' if got >= want else f' \033[33m✗\033[0m  '

        line = f'  R{r:<2d} ¥{rd[5]:>6.2f} {rd[11][:24]:24s} [{ctype:4s}] {cell("发票")}  {cell("行程单")}  {cell("付款截图")}  {cell("扫描件")}'
        print(line)
    print()

# ═══════════════════════════════════════════════════════════════════════
# 7. 文件覆盖自检
# ═══════════════════════════════════════════════════════════════════════

def _row_has_images(row_num, pdf_map, pay_map, scan_map):
    """某行是否至少有一张真实图片。空数组不算。"""
    return (
        len(pdf_map.get(row_num, [])) > 0
        or len(pay_map.get(row_num, [])) > 0
        or len(scan_map.get(row_num, [])) > 0
    )


def file_audit(rows, pdf_map, pay_map, scan_map, unmatched_pdfs):
    errors = []
    warnings = []

    data_rows = set(range(3, 3 + len(rows)))
    imaged = {r for r in data_rows if _row_has_images(r, pdf_map, pay_map, scan_map)}
    for r in sorted(data_rows - imaged):
        errors.append(f'R{r} 无任何图片')

    for f in unmatched_pdfs:
        errors.append(f'孤儿 PDF: {f}')

    pdf_in_use = {pf for sub in pdf_map.values() for pf in sub}
    for pf in sorted(pdf_in_use):
        if pf not in pdf_paths:
            errors.append(f'PDF 不存在: {pf}')

    # 付款截图/扫描件孤儿不阻止，仅警告
    for subdir, label, the_map in [
        ('付款截图', '付款截图', pay_map),
        ('扫描件', '扫描件', scan_map),
    ]:
        d = WORK / subdir
        on_disk = {f.name for f in d.glob('*') if f.suffix.lower() in ('.jpg', '.png', '.jpeg')} if d.exists() else set()
        in_use  = {pf for sub in the_map.values() for pf in sub}
        for f in sorted(on_disk - in_use):
            if not f.startswith('_'):  # skip temp/preview files
                warnings.append(f'孤儿 {label}: {f}')
        for f in sorted(in_use - on_disk):
            errors.append(f'{label} 不存在: {f}')

    if warnings:
        for w in warnings:
            print(f'  \033[33mWARN: {w}\033[0m')

    if errors:
        for e in errors:
            print(f'  ERROR: {e}')
        die('自检未通过，已阻止保存。')

    print(f'  self-check passed')
    print(f'  行: {len(imaged)}/{len(data_rows)}  '
          f'PDF: {len(pdf_in_use)}  '
          f'付款截图: {sum(len(v) for v in pay_map.values())}  '
          f'扫描件: {sum(len(v) for v in scan_map.values())}')

# ═══════════════════════════════════════════════════════════════════════
# 主流程
# ═══════════════════════════════════════════════════════════════════════

def die(msg):
    print(f'\n  \033[31m{msg}\033[0m')
    sys.exit(1)

def main():
    print(f'工作目录: {WORK}')

    # 1. 数据
    rows = load_csv()
    year_month = rows[0][6]
    for i, rd in enumerate(rows):
        if not re.match(r'^\d{4}-\d{2}$', rd[6]):
            print(f'  WARN: R{3+i} 月份格式异常 "{rd[6]}"，应为 YYYY-MM')
    print(f'  数据行: {len(rows)}  月份: {year_month}')
    data_rows = valid_data_rows(rows)

    # 2. PDF 自动匹配
    pdf_map, auto_unmatched, auto_unmapped = match_pdfs(rows)
    if auto_unmatched:
        print(f'  自动匹配未覆盖的 PDF: {auto_unmatched}')
    if auto_unmapped:
        print(f'  无 PDF 的行: R{auto_unmapped}')

    # 3. 加载并校验人工 PDF 映射
    manual_map, pdf_errors, pdf_warnings = load_pdf_mapping(
        'pdf_mapping.json', data_rows, pdf_paths,
        auto_pdf_map=pdf_map)

    # 4. 校验失败必须 fatal
    if pdf_errors:
        for e in pdf_errors:
            print(f'  \033[31mERROR: {e}\033[0m')
        die('pdf_mapping.json 校验未通过，已阻止保存。')

    # 5. 显示人工映射提示
    if pdf_warnings:
        for w in pdf_warnings:
            print(f'  \033[33mWARN: {w}\033[0m')

    # 6. 合并人工映射
    if manual_map:
        resolved = set()
        for r, pfs in manual_map.items():
            if not pfs:
                continue
            if r in pdf_map and pdf_map[r]:
                resolved.update(pf for pf in pfs)
            pdf_map[r] = pfs
            resolved.update(pf for pf in pfs if pf in auto_unmatched)
        if resolved:
            print(f'  已由人工映射解决: {sorted(resolved)}')
        print(f'  人工映射行: {sorted(r for r in manual_map if manual_map[r])}')

    # 7. 合并后重新计算真实未匹配状态
    unmatched_pdfs, unmapped_rows = recompute_pdf_status(rows, pdf_map)
    # 区分语义
    pending_review = sorted([pf for pf in unmatched_pdfs
                            if '滴滴' in pf or '高德' in pf])
    orphan_pdfs = sorted([pf for pf in unmatched_pdfs
                         if pf not in pending_review])

    if pending_review:
        print(f'  ⚠ 待人工确认 PDF ({len(pending_review)}): {pending_review}')
    if orphan_pdfs:
        print(f'  ⚠ 其他未匹配 PDF ({len(orphan_pdfs)}): {orphan_pdfs}')
    if unmapped_rows:
        print(f'  无 PDF 的行: R{unmapped_rows}')
    if not unmatched_pdfs and not unmapped_rows:
        print('  ✓ 所有 PDF 已匹配，所有行已有 PDF')

    # 8. 刷新 review 文件（基于合并后状态）
    write_pdf_review_if_needed(rows, pdf_map)

    # 9. 加载付款截图和扫描件映射
    pay_map = {}
    pay_path = WORK / 'payment_mapping.json'
    if pay_path.exists():
        with open(pay_path, encoding='utf-8') as f:
            raw = json.load(f)
            pay_map = {int(k): v for k, v in raw.items()}

    scan_map = {}
    scan_path = WORK / 'scan_mapping.json'
    if scan_path.exists():
        with open(scan_path, encoding='utf-8') as f:
            raw = json.load(f)
            scan_map = {int(k): v for k, v in raw.items()}

    # 10. 文档检查
    check_completeness(rows, pdf_map, pay_map, scan_map)

    # 11. 加载模板
    if not TEMPLATE.exists():
        die(f'模板不存在: {TEMPLATE}')
    wb = openpyxl.load_workbook(str(TEMPLATE))
    ws = wb.active

    for mc in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mc))
    for r in range(3, 50):
        for c in range(1, 13):
            ws.cell(row=r, column=c).value = None
    ws._images = []
    ws._drawing = None

    # 12. 填数据
    fill_sheet(ws, rows)

    # 13. 嵌图片
    (WORK / '付款截图').mkdir(exist_ok=True)
    (WORK / '扫描件').mkdir(exist_ok=True)
    embed_images(ws, pdf_map, pay_map, scan_map)

    # 14. 自检
    file_audit(rows, pdf_map, pay_map, scan_map, unmatched_pdfs)

    # 15. 保存
    dst = WORK / f'报销-{year_month[2:4]}-{year_month[5:]}.xlsx'
    try:
        wb.save(str(dst))
        print(f'\nSaved: {dst}')
    except PermissionError:
        fallback = WORK / f'报销-{year_month[2:4]}-{year_month[5:]}-NEW.xlsx'
        wb.save(str(fallback))
        print(f'\nTarget locked. Saved: {fallback}')

if __name__ == '__main__':
    main()
